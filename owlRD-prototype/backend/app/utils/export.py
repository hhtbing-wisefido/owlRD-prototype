"""
数据导出工具
支持导出为CSV、Excel、JSON格式
"""

import csv
import json
from typing import List, Dict, Any
from io import StringIO, BytesIO
from datetime import datetime


def export_to_csv(data: List[Dict[str, Any]], columns: List[str] = None) -> str:
    """
    导出数据为CSV格式
    
    Args:
        data: 数据列表
        columns: 导出的列（如果为None，使用第一条数据的所有键）
    
    Returns:
        CSV字符串
    """
    if not data:
        return ""
    
    # 确定列
    if columns is None:
        columns = list(data[0].keys())
    
    # 创建CSV
    output = StringIO()
    writer = csv.DictWriter(output, fieldnames=columns, extrasaction='ignore')
    
    # 写入表头
    writer.writeheader()
    
    # 写入数据
    for row in data:
        # 转换复杂类型为字符串
        processed_row = {}
        for col in columns:
            value = row.get(col, '')
            if isinstance(value, (list, dict)):
                processed_row[col] = json.dumps(value, ensure_ascii=False)
            else:
                processed_row[col] = value
        writer.writerow(processed_row)
    
    return output.getvalue()


def export_to_json(data: List[Dict[str, Any]], pretty: bool = True) -> str:
    """
    导出数据为JSON格式
    
    Args:
        data: 数据列表
        pretty: 是否美化输出
    
    Returns:
        JSON字符串
    """
    if pretty:
        return json.dumps(data, ensure_ascii=False, indent=2)
    return json.dumps(data, ensure_ascii=False)


def export_to_excel(data: List[Dict[str, Any]], columns: List[str] = None) -> bytes:
    """
    导出数据为Excel格式 (需要openpyxl库)
    
    Args:
        data: 数据列表
        columns: 导出的列
    
    Returns:
        Excel文件字节流
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError:
        raise ImportError("需要安装openpyxl库: pip install openpyxl")
    
    if not data:
        return b""
    
    # 确定列
    if columns is None:
        columns = list(data[0].keys())
    
    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "数据导出"
    
    # 样式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # 写入表头
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # 写入数据
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, col_name in enumerate(columns, 1):
            value = row_data.get(col_name, '')
            # 转换复杂类型
            if isinstance(value, (list, dict)):
                value = json.dumps(value, ensure_ascii=False)
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # 自动调整列宽
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # 保存到BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def generate_filename(prefix: str, extension: str) -> str:
    """
    生成导出文件名
    
    Args:
        prefix: 文件名前缀
        extension: 文件扩展名（不含点）
    
    Returns:
        文件名
    """
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"{prefix}_{timestamp}.{extension}"


def filter_columns(
    data: List[Dict[str, Any]],
    include: List[str] = None,
    exclude: List[str] = None
) -> List[Dict[str, Any]]:
    """
    过滤数据列
    
    Args:
        data: 原始数据
        include: 包含的列（如果指定，只导出这些列）
        exclude: 排除的列
    
    Returns:
        过滤后的数据
    """
    if not data:
        return []
    
    filtered_data = []
    
    for row in data:
        filtered_row = {}
        
        for key, value in row.items():
            # 应用include过滤
            if include and key not in include:
                continue
            
            # 应用exclude过滤
            if exclude and key in exclude:
                continue
            
            filtered_row[key] = value
        
        filtered_data.append(filtered_row)
    
    return filtered_data


# 预定义的列配置
EXPORT_COLUMNS = {
    "users": [
        "user_id", "username", "email", "phone", "role",
        "is_active", "created_at"
    ],
    "residents": [
        "resident_id", "last_name", "first_name", "gender",
        "date_of_birth", "admission_date", "status", "room_number"
    ],
    "devices": [
        "device_id", "device_code", "device_type", "mac_address",
        "location_id", "status", "installed", "monitoring_enabled"
    ],
    "alerts": [
        "alert_id", "alert_type", "severity", "resident_id",
        "location_id", "alert_time", "status", "response_time"
    ]
}


def get_export_columns(entity_type: str) -> List[str]:
    """
    获取实体类型的导出列配置
    
    Args:
        entity_type: 实体类型
    
    Returns:
        列名列表
    """
    return EXPORT_COLUMNS.get(entity_type, [])
